from io import BytesIO

from django.db import transaction
from django.http import HttpResponse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_framework_simplejwt.tokens import RefreshToken

from .models import User
from .permissions import IsSuperAdmin
from .serializers import (
    AdminUserUpdateSerializer,
    BulkUserImportSerializer,
    ChangePasswordSerializer,
    CreateUserSerializer,
    FirstPasswordChangeSerializer,
    LoginSerializer,
    LogoutSerializer,
    ProfileSerializer,
    UserListSerializer,
)


# =========================================================
# BULK IMPORT CONSTANTS
# =========================================================


BULK_USER_HEADERS = [
    "role",
    "student_id",
    "email",
    "first_name",
    "last_name",
    "phone",
    "password",
    "department",
    "batch",
    "semester",
    "designation",
    "project_start_term",
    "project_start_year",
    "project_duration",
]


ALLOWED_BULK_ROLES = {
    "STUDENT",
    "SUPERVISOR",
    "EXAMINER",
}


# =========================================================
# COMMON HELPERS
# =========================================================


def normalize_excel_header(value):
    """
    Convert an Excel header into a normalized API field name.
    """

    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )


def normalize_excel_text(value):
    """
    Convert Excel values into clean strings.

    Integer-looking float values, such as 223001.0, become 223001.
    """

    if value is None:
        return ""

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def normalize_optional_integer(value, field_name):
    """
    Convert an optional Excel value to an integer.
    """

    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, bool):
        raise ValueError(
            f"{field_name} must be a valid integer."
        )

    try:
        numeric_value = float(value)

        if not numeric_value.is_integer():
            raise ValueError

        return int(numeric_value)

    except (TypeError, ValueError):
        raise ValueError(
            f"{field_name} must be a valid integer."
        )


def flatten_serializer_errors(errors):
    """
    Convert nested DRF serializer errors into a readable string.
    """

    messages = []

    def collect(value, prefix=""):
        if isinstance(value, dict):
            for key, child_value in value.items():
                child_prefix = (
                    f"{prefix}.{key}"
                    if prefix
                    else str(key)
                )

                collect(
                    child_value,
                    child_prefix,
                )

        elif isinstance(value, (list, tuple)):
            for child_value in value:
                collect(
                    child_value,
                    prefix,
                )

        else:
            text = str(value)

            if prefix:
                messages.append(
                    f"{prefix}: {text}"
                )
            else:
                messages.append(text)

    collect(errors)

    return " | ".join(messages) or "Invalid row data."


def row_is_empty(row_values):
    """
    Return True when every cell in an Excel row is empty.
    """

    return all(
        value is None
        or str(value).strip() == ""
        for value in row_values
    )


def prepare_bulk_user_payload(row_data):
    """
    Normalize and validate one Excel row before passing it to
    CreateUserSerializer.
    """

    role = normalize_excel_text(
        row_data.get("role")
    ).upper()

    student_id = normalize_excel_text(
        row_data.get("student_id")
    )

    email = normalize_excel_text(
        row_data.get("email")
    ).lower()

    first_name = normalize_excel_text(
        row_data.get("first_name")
    )

    last_name = normalize_excel_text(
        row_data.get("last_name")
    )

    phone = normalize_excel_text(
        row_data.get("phone")
    )

    password = normalize_excel_text(
        row_data.get("password")
    )

    department = normalize_excel_text(
        row_data.get("department")
    )

    batch = normalize_excel_text(
        row_data.get("batch")
    )

    semester = normalize_excel_text(
        row_data.get("semester")
    )

    designation = normalize_excel_text(
        row_data.get("designation")
    )

    project_start_term = normalize_excel_text(
        row_data.get("project_start_term")
    ).upper()

    project_start_year = normalize_optional_integer(
        row_data.get("project_start_year"),
        "Project start year",
    )

    project_duration = normalize_optional_integer(
        row_data.get("project_duration"),
        "Project duration",
    )

    if not role:
        raise ValueError("Role is required.")

    if role not in ALLOWED_BULK_ROLES:
        raise ValueError(
            "Role must be STUDENT, SUPERVISOR or EXAMINER."
        )

    if not first_name:
        raise ValueError("First name is required.")

    if not password:
        raise ValueError("Password is required.")

    if len(password) < 8:
        raise ValueError(
            "Password must contain at least 8 characters."
        )

    payload = {
        "role": role,
        "student_id": student_id or None,
        "email": email or None,
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "password": password,
        "department": department,
        "batch": batch,
        "semester": semester,
        "designation": designation,
    }

    if role == "STUDENT":
        if not student_id:
            raise ValueError(
                "Student ID is required for a student."
            )

        if not project_start_term:
            raise ValueError(
                "Project start semester is required for a student."
            )

        if project_start_year is None:
            raise ValueError(
                "Project start year is required for a student."
            )

        payload["email"] = None
        payload["project_start_term"] = project_start_term
        payload["project_start_year"] = project_start_year
        payload["project_duration"] = (
            project_duration
            if project_duration is not None
            else 3
        )

    else:
        if not email:
            raise ValueError(
                "Email is required for a supervisor or examiner."
            )

        payload["student_id"] = None

        # Semester fields are not used by Supervisor or Examiner.
        payload["project_start_term"] = None
        payload["project_start_year"] = None
        payload["project_duration"] = 3

    return payload


def get_user_identifier(payload):
    """
    Return the most useful identifier for an import result.
    """

    if payload.get("role") == "STUDENT":
        return payload.get("student_id") or "-"

    return payload.get("email") or "-"


# =========================================================
# AUTHENTICATION
# =========================================================


class LoginAPIView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        serializer = LoginSerializer(
            data=request.data,
        )

        serializer.is_valid(
            raise_exception=True,
        )

        user = serializer.validated_data["user"]

        refresh = RefreshToken.for_user(user)

        return Response(
            {
                "success": True,
                "message": "Login successful.",
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "force_password_change": (
                    user.must_change_password
                ),
                "user": ProfileSerializer(
                    user,
                    context={"request": request},
                ).data,
            },
            status=status.HTTP_200_OK,
        )


class ProfileAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
    ]

    def get(self, request):
        serializer = ProfileSerializer(
            request.user,
            context={"request": request},
        )

        return Response({
            "success": True,
            "data": serializer.data,
        })


class FirstPasswordChangeAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
    ]

    def post(self, request):
        serializer = FirstPasswordChangeSerializer(
            data=request.data,
        )

        serializer.is_valid(
            raise_exception=True,
        )

        user = request.user

        user.set_password(
            serializer.validated_data[
                "new_password"
            ]
        )

        user.is_first_login = False
        user.must_change_password = False

        user.save(
            update_fields=[
                "password",
                "is_first_login",
                "must_change_password",
                "updated_at",
            ]
        )

        return Response({
            "success": True,
            "message": (
                "Password changed successfully. "
                "You can now access the dashboard."
            ),
        })


class ChangePasswordAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
    ]

    def post(self, request):
        serializer = ChangePasswordSerializer(
            data=request.data,
            context={"request": request},
        )

        serializer.is_valid(
            raise_exception=True,
        )

        user = request.user

        user.set_password(
            serializer.validated_data[
                "new_password"
            ]
        )

        user.save(
            update_fields=[
                "password",
                "updated_at",
            ]
        )

        return Response({
            "success": True,
            "message": (
                "Password changed successfully."
            ),
        })


class LogoutAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
    ]

    def post(self, request):
        serializer = LogoutSerializer(
            data=request.data,
        )

        serializer.is_valid(
            raise_exception=True,
        )

        serializer.save()

        return Response(
            {
                "success": True,
                "message": "Logout successful.",
            },
            status=status.HTTP_200_OK,
        )


# =========================================================
# SINGLE USER CREATE
# =========================================================


class CreateUserAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def post(self, request):
        serializer = CreateUserSerializer(
            data=request.data,
            context={"request": request},
        )

        serializer.is_valid(
            raise_exception=True,
        )

        user = serializer.save()

        response_serializer = UserListSerializer(
            user,
            context={"request": request},
        )

        return Response(
            {
                "success": True,
                "message": (
                    "User created successfully."
                ),
                "data": response_serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )


# =========================================================
# BULK USER IMPORT
# =========================================================


class BulkUserImportAPIView(APIView):
    """
    Import Student, Supervisor and Examiner accounts from an
    .xlsx Excel file.

    The complete file is validated first. If any row is invalid,
    no user is created. All valid rows are then created inside
    one database transaction.
    """

    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    parser_classes = [
        MultiPartParser,
        FormParser,
    ]

    def post(self, request):
        upload_serializer = BulkUserImportSerializer(
            data=request.data,
        )

        upload_serializer.is_valid(
            raise_exception=True,
        )

        uploaded_file = upload_serializer.validated_data[
            "file"
        ]

        try:
            workbook = load_workbook(
                filename=uploaded_file,
                read_only=True,
                data_only=True,
            )

        except Exception:
            return Response(
                {
                    "success": False,
                    "message": (
                        "The Excel file could not be read. "
                        "Please upload a valid .xlsx file."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            worksheet = workbook.active

            rows = worksheet.iter_rows(
                values_only=True,
            )

            try:
                raw_headers = next(rows)

            except StopIteration:
                return Response(
                    {
                        "success": False,
                        "message": (
                            "The uploaded Excel file is empty."
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            normalized_headers = [
                normalize_excel_header(header)
                for header in raw_headers
            ]

            duplicate_headers = {
                header
                for header in normalized_headers
                if header
                and normalized_headers.count(header) > 1
            }

            if duplicate_headers:
                return Response(
                    {
                        "success": False,
                        "message": (
                            "Duplicate Excel headers found: "
                            + ", ".join(
                                sorted(duplicate_headers)
                            )
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            missing_headers = [
                header
                for header in BULK_USER_HEADERS
                if header not in normalized_headers
            ]

            if missing_headers:
                return Response(
                    {
                        "success": False,
                        "message": (
                            "Required Excel column(s) are missing: "
                            + ", ".join(missing_headers)
                        ),
                        "required_headers": BULK_USER_HEADERS,
                        "received_headers": normalized_headers,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            header_indexes = {
                header: normalized_headers.index(header)
                for header in BULK_USER_HEADERS
            }

            prepared_rows = []
            errors = []
            total_rows = 0

            excel_student_ids = set()
            excel_emails = set()

            for excel_row_number, row_values in enumerate(
                rows,
                start=2,
            ):
                if row_is_empty(row_values):
                    continue

                total_rows += 1

                row_data = {}

                for header, column_index in header_indexes.items():
                    row_data[header] = (
                        row_values[column_index]
                        if column_index < len(row_values)
                        else None
                    )

                raw_identifier = (
                    normalize_excel_text(
                        row_data.get("student_id")
                    )
                    or normalize_excel_text(
                        row_data.get("email")
                    )
                    or "-"
                )

                try:
                    payload = prepare_bulk_user_payload(
                        row_data
                    )

                    identifier = get_user_identifier(
                        payload
                    )

                    if payload["role"] == "STUDENT":
                        student_id_key = (
                            payload["student_id"]
                            .strip()
                            .lower()
                        )

                        if student_id_key in excel_student_ids:
                            raise ValueError(
                                "Duplicate student ID found "
                                "inside the Excel file."
                            )

                        excel_student_ids.add(
                            student_id_key
                        )

                    else:
                        email_key = (
                            payload["email"]
                            .strip()
                            .lower()
                        )

                        if email_key in excel_emails:
                            raise ValueError(
                                "Duplicate email found "
                                "inside the Excel file."
                            )

                        excel_emails.add(
                            email_key
                        )

                    serializer = CreateUserSerializer(
                        data=payload,
                        context={
                            "request": request,
                        },
                    )

                    serializer.is_valid(
                        raise_exception=True,
                    )

                    prepared_rows.append({
                        "row": excel_row_number,
                        "identifier": identifier,
                        "serializer": serializer,
                    })

                except ValueError as error:
                    errors.append({
                        "row": excel_row_number,
                        "identifier": raw_identifier,
                        "message": str(error),
                    })

                except Exception as error:
                    error_detail = getattr(
                        error,
                        "detail",
                        None,
                    )

                    if error_detail is not None:
                        message = flatten_serializer_errors(
                            error_detail
                        )
                    else:
                        message = str(error)

                    errors.append({
                        "row": excel_row_number,
                        "identifier": raw_identifier,
                        "message": (
                            message
                            or "Row validation failed."
                        ),
                    })

            if total_rows == 0:
                return Response(
                    {
                        "success": False,
                        "message": (
                            "No user data rows were found "
                            "in the Excel file."
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Stop before creating anything when any row fails.
            if errors:
                return Response(
                    {
                        "success": False,
                        "message": (
                            f"Bulk import cancelled. "
                            f"{len(errors)} row(s) contain errors. "
                            f"No user was created."
                        ),
                        "total_rows": total_rows,
                        "created_count": 0,
                        "failed_count": len(errors),
                        "created_users": [],
                        "errors": errors,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            created_users = []

            # All rows are created in one transaction.
            # An unexpected failure rolls back the complete import.
            try:
                with transaction.atomic():
                    for prepared_row in prepared_rows:
                        serializer = prepared_row[
                            "serializer"
                        ]

                        user = serializer.save()

                        created_users.append({
                            "row": prepared_row["row"],
                            "id": user.id,
                            "role": user.role,
                            "identifier": (
                                user.student_id
                                or user.email
                                or "-"
                            ),
                            "name": (
                                f"{user.first_name or ''} "
                                f"{user.last_name or ''}"
                            ).strip(),
                        })

            except Exception as error:
                return Response(
                    {
                        "success": False,
                        "message": (
                            "Bulk import failed while creating "
                            "users. No user was saved."
                        ),
                        "detail": str(error),
                        "total_rows": total_rows,
                        "created_count": 0,
                        "failed_count": total_rows,
                        "created_users": [],
                        "errors": [],
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            return Response(
                {
                    "success": True,
                    "message": (
                        f"Bulk import completed successfully. "
                        f"{len(created_users)} user(s) created."
                    ),
                    "total_rows": total_rows,
                    "created_count": len(created_users),
                    "failed_count": 0,
                    "created_users": created_users,
                    "errors": [],
                },
                status=status.HTTP_201_CREATED,
            )

        finally:
            workbook.close()

# =========================================================
# BULK IMPORT TEMPLATE DOWNLOAD
# =========================================================


class BulkUserTemplateDownloadAPIView(APIView):
    """
    Download a ready-to-use Excel template for bulk user import.
    """

    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def get(self, request):
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Users"

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1E3A5F",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        required_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        example_fill = PatternFill(
            fill_type="solid",
            fgColor="F2F2F2",
        )

        for column_index, header in enumerate(
            BULK_USER_HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=header,
            )

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        example_rows = [
            [
                "STUDENT",
                "223001",
                "",
                "Rahim",
                "Ahmed",
                "01700000001",
                "Student@123",
                "CSE",
                "223",
                "8th",
                "",
                "FALL",
                2026,
                3,
            ],
            [
                "SUPERVISOR",
                "",
                "supervisor@example.com",
                "Karim",
                "Hasan",
                "01800000001",
                "Teacher@123",
                "CSE",
                "",
                "",
                "Lecturer",
                "",
                "",
                "",
            ],
            [
                "EXAMINER",
                "",
                "examiner@example.com",
                "Salma",
                "Akter",
                "01900000001",
                "Examiner@123",
                "CSE",
                "",
                "",
                "Associate Professor",
                "",
                "",
                "",
            ],
        ]

        for row_index, example_row in enumerate(
            example_rows,
            start=2,
        ):
            for column_index, value in enumerate(
                example_row,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

                cell.fill = example_fill
                cell.alignment = Alignment(
                    vertical="top",
                )

        # Keep IDs, phone numbers and passwords as text.
        text_columns = {
            "student_id",
            "phone",
            "password",
            "batch",
            "semester",
        }

        for header in text_columns:
            column_index = (
                BULK_USER_HEADERS.index(header)
                + 1
            )

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_index,
                ).number_format = "@"

        required_headers = {
            "role",
            "first_name",
            "password",
        }

        for header in required_headers:
            column_index = (
                BULK_USER_HEADERS.index(header)
                + 1
            )

            worksheet.cell(
                row=1,
                column=column_index,
            ).fill = required_fill

            worksheet.cell(
                row=1,
                column=column_index,
            ).font = Font(
                color="000000",
                bold=True,
            )

        column_widths = {
            "role": 18,
            "student_id": 18,
            "email": 30,
            "first_name": 18,
            "last_name": 18,
            "phone": 18,
            "password": 20,
            "department": 18,
            "batch": 14,
            "semester": 14,
            "designation": 24,
            "project_start_term": 24,
            "project_start_year": 22,
            "project_duration": 20,
        }

        for column_index, header in enumerate(
            BULK_USER_HEADERS,
            start=1,
        ):
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = column_widths.get(
                header,
                18,
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(BULK_USER_HEADERS))}"
            f"{worksheet.max_row}"
        )

        instruction_sheet = (
            workbook.create_sheet(
                title="Instructions"
            )
        )

        instructions = [
            [
                "Bulk User Import Instructions"
            ],
            [
                "1. Do not rename or delete any header column."
            ],
            [
                "2. Allowed roles: STUDENT, SUPERVISOR, EXAMINER."
            ],
            [
                "3. Every row requires role, first_name and password."
            ],
            [
                "4. Password must contain at least 8 characters."
            ],
            [
                (
                    "5. A STUDENT requires student_id, "
                    "project_start_term and project_start_year."
                )
            ],
            [
                (
                    "6. Allowed project_start_term values: "
                    "SPRING, SUMMER, FALL."
                )
            ],
            [
                (
                    "7. A SUPERVISOR or EXAMINER requires email."
                )
            ],
            [
                (
                    "8. project_duration defaults to 3 when "
                    "left blank for a student."
                )
            ],
            [
                (
                    "9. Keep Student ID and phone cells formatted "
                    "as Text to preserve leading zeroes."
                )
            ],
            [
                (
                    "10. Remove the example rows before importing "
                    "your real users."
                )
            ],
        ]

        for row_index, instruction in enumerate(
            instructions,
            start=1,
        ):
            cell = instruction_sheet.cell(
                row=row_index,
                column=1,
                value=instruction[0],
            )

            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

            if row_index == 1:
                cell.font = Font(
                    bold=True,
                    size=14,
                )

        instruction_sheet.column_dimensions[
            "A"
        ].width = 95

        output = BytesIO()

        workbook.save(output)
        workbook.close()

        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        response[
            "Content-Disposition"
        ] = (
            'attachment; filename="bulk_user_import_template.xlsx"'
        )

        return response


# =========================================================
# USER LIST AND DETAILS
# =========================================================


class UserListAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def get(self, request):
        users = (
            User.objects.select_related(
                "created_by",
            )
            .all()
            .order_by("-id")
        )

        serializer = UserListSerializer(
            users,
            many=True,
            context={"request": request},
        )

        return Response({
            "success": True,
            "count": users.count(),
            "data": serializer.data,
        })


class UserDetailAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def get(self, request, user_id):
        try:
            user = User.objects.get(
                id=user_id,
            )

        except User.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "User not found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = UserListSerializer(
            user,
            context={"request": request},
        )

        return Response({
            "success": True,
            "data": serializer.data,
        })


# =========================================================
# USER STATUS UPDATE
# =========================================================


class UserStatusUpdateAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def patch(self, request, user_id):
        try:
            user = User.objects.get(
                id=user_id,
            )

        except User.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "User not found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        if user.role == "SUPER_ADMIN":
            return Response(
                {
                    "success": False,
                    "message": (
                        "Super Admin status cannot be changed."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        is_active = request.data.get(
            "is_active"
        )

        if not isinstance(is_active, bool):
            return Response(
                {
                    "success": False,
                    "message": (
                        "is_active must be true or false."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.is_active = is_active

        user.save(
            update_fields=[
                "is_active",
                "updated_at",
            ]
        )

        return Response({
            "success": True,
            "message": (
                "User status updated successfully."
            ),
            "data": {
                "id": user.id,
                "email": user.email,
                "student_id": user.student_id,
                "role": user.role,
                "is_active": user.is_active,
            },
        })


# =========================================================
# ADMIN USER UPDATE
# =========================================================


class AdminUserUpdateAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def put(self, request, user_id):
        return self.update_user(
            request,
            user_id,
        )

    def patch(self, request, user_id):
        return self.update_user(
            request,
            user_id,
        )

    def update_user(self, request, user_id):
        try:
            user = User.objects.get(
                id=user_id,
            )

        except User.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "User not found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = AdminUserUpdateSerializer(
            user,
            data=request.data,
            partial=True,
            context={"request": request},
        )

        serializer.is_valid(
            raise_exception=True,
        )

        updated_user = serializer.save()

        response_serializer = UserListSerializer(
            updated_user,
            context={"request": request},
        )

        return Response({
            "success": True,
            "message": (
                "User updated successfully."
            ),
            "data": response_serializer.data,
        })


# =========================================================
# ADMIN USER DELETE
# =========================================================


class AdminUserDeleteAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def delete(self, request, user_id):
        try:
            user = User.objects.get(
                id=user_id,
            )

        except User.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "User not found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        if user.id == request.user.id:
            return Response(
                {
                    "success": False,
                    "message": (
                        "You cannot delete your own account."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if user.role == "SUPER_ADMIN":
            return Response(
                {
                    "success": False,
                    "message": (
                        "A Super Admin account cannot be "
                        "deleted from this endpoint."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.delete()

        return Response({
            "success": True,
            "message": (
                "User deleted successfully."
            ),
        })
        
        
        
class BulkUserDeleteAPIView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsSuperAdmin,
    ]

    def post(self, request):
        user_ids = request.data.get("user_ids", [])

        if not isinstance(user_ids, list):
            return Response(
                {
                    "success": False,
                    "message": "user_ids must be a list.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user_ids = [
                int(user_id)
                for user_id in user_ids
            ]
        except (TypeError, ValueError):
            return Response(
                {
                    "success": False,
                    "message": "One or more user IDs are invalid.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        user_ids = list(set(user_ids))

        if not user_ids:
            return Response(
                {
                    "success": False,
                    "message": "Please select at least one user.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if request.user.id in user_ids:
            return Response(
                {
                    "success": False,
                    "message": (
                        "You cannot delete your own account."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        users = User.objects.filter(
            id__in=user_ids,
        )

        super_admin_exists = users.filter(
            role="SUPER_ADMIN",
        ).exists()

        if super_admin_exists:
            return Response(
                {
                    "success": False,
                    "message": (
                        "Super Admin accounts cannot be deleted."
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        found_ids = set(
            users.values_list(
                "id",
                flat=True,
            )
        )

        missing_ids = [
            user_id
            for user_id in user_ids
            if user_id not in found_ids
        ]

        deleted_count = users.count()

        with transaction.atomic():
            users.delete()

        return Response(
            {
                "success": True,
                "message": (
                    f"{deleted_count} user(s) deleted successfully."
                ),
                "deleted_count": deleted_count,
                "missing_ids": missing_ids,
            },
            status=status.HTTP_200_OK,
        )